
import tkinter as tk
import csv
from tkinter import filedialog,ttk, messagebox
import sqlite3
from utils import resource_path, validate_date, validate_nni, convert_dmy_to_ymd, convert_ymd_to_dmy
from utils import show_etudiant_profile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

DB_PATH = "identifier.sqlite"

def recherche(entries):
    """" recherche dans la BDD avec les criteres """
    values = {}
    for field, entry in entries.items():
        try:
            values[field] = entry.get().strip()
        except AttributeError:
            print(f"Champ '{field}' n’est pas un Entry valide : {entry}")

    conditions = []
    parameters = []

    if values["nom"]:
        conditions.append("nom LIKE ?")
        parameters.append(f"%{values['nom']}%")
    if values["prenom"]:
        conditions.append("prenom LIKE ?")
        parameters.append(f"%{values['prenom']}%")
    if values["nni"]:
        if not validate_nni(values["nni"]):
            messagebox.showerror("Erreur", "NNI invalide.")
            return
        conditions.append("nni = ?")
        parameters.append(values["nni"])
    if values["dob"]:
        try:
            date_sql = convert_dmy_to_ymd(values["dob"])
            if not validate_date(date_sql):
                raise ValueError
            conditions.append("dob = ?")
            parameters.append(date_sql)
        except:
            messagebox.showerror("Erreur", "Format de date invalide.")
            return
    if values["place_of_birth"]:
        conditions.append("place_of_birth LIKE ?")
        parameters.append(f"%{values['place_of_birth']}%")
    if values["field_of_study"]:
        conditions.append("field_of_study LIKE ?")
        parameters.append(f"%{values['field_of_study']}%")
    if values["academic_year"]:
        conditions.append("academic_year LIKE ?")
        parameters.append(f"%{values['academic_year']}%")
    if values["type_convention"]:
        conditions.append("type_convention = ?")
        parameters.append(values["type_convention"])
    if values["study_location"]:
        conditions.append("study_location LIKE ?")
        parameters.append(f"%{values['study_location']}%")
    if values["email"]:
        conditions.append("email LIKE ?")
        parameters.append(f"%{values['email']}%")
    if values["phone_number"]:
        conditions.append("phone_number LIKE ?")
        parameters.append(f"%{values['phone_number']}%")

    query = "SELECT * FROM T_etudiant_etd"
    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute(query, parameters)
        results = cursor.fetchall()

        if not results:
            messagebox.showinfo("Résultat", "Aucun étudiant trouvé.")
        else:
            result_win = tk.Toplevel()
            result_win.title("Résultats de la recherche")
            result_win.geometry("1000x400")

            tree = ttk.Treeview(result_win)
            tree.pack(fill="both", expand=True)

            cursor.execute("PRAGMA table_info(T_etudiant_etd)")
            columns = [info[1] for info in cursor.fetchall()]

            tree["columns"] = columns
            tree["show"] = "headings"
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=100)

            for row in results:
                tree.insert("", "end", values=row)

            def on_double_click(event):
                selected_item = tree.selection()
                if selected_item:
                    item_data = tree.item(selected_item)
                    values = item_data["values"]
                    etudiant_id = values[0]
                    show_etudiant_profile(etudiant_id, lambda: recherche(entries))

            tree.bind("<Double-1>", on_double_click)
            ttk.Button(result_win, text="Exporter sélection personnalisée",
                       command=lambda: ask_and_export_csv_with_bank(tree)).pack(pady=10)

            # Nouveau bouton pour exporter au format BOURSE
            ttk.Button(result_win, text="Exporter au format BOURSE",
                       command=lambda: export_bourse_modele(tree)).pack(pady=10)

        conn.close()
    except Exception as e:
        messagebox.showerror("Erreur", f"Une erreur est survenue : {e}")

def ask_and_export_csv_with_bank(tree):
    columns = tree["columns"]

    column_labels = {
        "T_etudiant_etd_pk": "ID",
        "dob": "Date de naissance",
        "nni": "NNI",
        "type_convention": "Convention",
        "place_of_birth": "Lieu de naissance",
        "study_location": "Ville d'études",
        "field_of_study": "Filière",
        "academic_year": "Niveau",
        "nom": "Nom",
        "prenom": "Prénom",
        "email": "Email",
        "phone_number": "Téléphone",
        "iban": "IBAN",
        "bic": "BIC",
        "montant": "Montant"
    }

    win = tk.Toplevel()
    win.title("Exporter vers CSV / Excel")
    win.geometry("650x700")

    vars = {}
    entries = {}

    label_frame = ttk.LabelFrame(win, text="Colonnes à exporter")
    label_frame.pack(padx=10, pady=10, fill="both", expand=True)

    def toggle_all(state=True):
        for var in vars.values():
            var.set(state)

    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=5)
    tk.Button(btn_frame, text="Tout sélectionner", command=lambda: toggle_all(True)).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Tout désélectionner", command=lambda: toggle_all(False)).pack(side="left", padx=5)

    for i, col in enumerate(columns):
        var = tk.BooleanVar(value=True)
        label = column_labels.get(col, col)
        tk.Checkbutton(label_frame, text=label, variable=var).grid(row=i, column=0, sticky="w", padx=10, pady=3)
        entry = tk.Entry(label_frame)
        entry.insert(0, label)
        entry.grid(row=i, column=1, padx=10, pady=3)
        vars[col] = var
        entries[col] = entry

    format_var = tk.StringVar(value="CSV")
    tk.Label(win, text="Format d'export :").pack()
    ttk.Combobox(win, textvariable=format_var, values=["CSV", "Excel"], state="readonly").pack(pady=5)

    def export():
        selected_items = tree.selection()
        if not selected_items:
            selected_items = tree.get_children()

        selected_columns = [col for col in vars if vars[col].get()]
        if not selected_columns:
            messagebox.showwarning("Erreur", "Aucune colonne sélectionnée.")
            return

        renames = {col: entries[col].get().strip() or col for col in selected_columns}
        filetype = format_var.get()

        filetypes = [("Fichiers CSV", "*.csv")] if filetype == "CSV" else [("Fichiers Excel", "*.xlsx")]
        default_ext = ".csv" if filetype == "CSV" else ".xlsx"

        file_path = filedialog.asksaveasfilename(
            defaultextension=default_ext,
            filetypes=filetypes,
            title="Enregistrer sous..."
        )
        if not file_path:
            return

        try:
            rows = []

            for item in selected_items:
                values = tree.item(item)["values"]
                row_data = {col: values[idx] for idx, col in enumerate(columns)}

                formatted_row = []
                for col in selected_columns:
                    val = row_data.get(col, "")
                    if val in [None, "None"]:
                        val = ""
                    if col == "dob" and val:
                        val = convert_ymd_to_dmy(val)
                    formatted_row.append(val)

                rows.append(formatted_row)

            if filetype == "CSV":
                with open(file_path, mode="w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow([renames[col] for col in selected_columns])
                    writer.writerows(rows)
            else:
                wb = Workbook()
                ws = wb.active
                ws.append([renames[col] for col in selected_columns])
                for row in rows:
                    ws.append(row)
                wb.save(file_path)

            messagebox.showinfo("Succès", f"Fichier exporté avec succès :\n{file_path}")
            win.destroy()

        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur d'export : {e}")

    tk.Button(win, text="Exporter", command=export).pack(pady=10)

def format_montant(value):
    """Formate un montant pour Excel FR (avec virgule) ou retourne 0 si vide"""
    try:
        val = float(value)
        return str(val).replace(".", ",")
    except:
        return "0"

def format_cell(value):
    """Remplace None ou 'None' par une chaîne vide"""
    return "" if value in [None, "None"] else str(value)


def export_bourse_modele(tree):
    columns = tree["columns"]
    selected_items = tree.selection() or tree.get_children()

    selected_data = []
    for item in selected_items:
        values = tree.item(item)["values"]
        row_data = {col: values[idx] if idx < len(values) else "" for idx, col in enumerate(columns)}
        selected_data.append(row_data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Feuil1"

    ws["B1"] = "AMBASSADE DE MAURITANIE A PARIS"; ws.merge_cells("B1:E1")
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws["B2"] = "Etat des Bourses 01/06/2025 AU 31/09/2025"; ws.merge_cells("B2:E2")
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=2).font = Font(bold=True)
    ws["B3"] = "Virement à effectuer sur le compte"; ws.merge_cells("B3:E3")
    ws.cell(row=3, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=2).font = Font(bold=True)
    ws["B4"] = ""; ws.merge_cells("B4:E4")
    ws["B5"] = "IBAN: FR76 3000 3016 7800 0500 2367 017 et CODE BIC: SOGEFRPP"; ws.merge_cells("B5:E5")
    ws.cell(row=5, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=5, column=2).font = Font(bold=True)
    ws.append([]); ws.append([])

    headers = ["N°", "NOMS & PRÉNOMS", "MONTANT", "CODE BIC", "CODE IBAN"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=8, column=col).font = Font(bold=True)

    for idx, row in enumerate(selected_data, start=1):
        full_name = f"{row.get('nom', '')} {row.get('prenom', '')}".strip()
        try:
            montant = float(row.get("montant", 0))
        except:
            montant = 0
        bic = format_cell(row.get("bic", ""))
        iban = format_cell(row.get("iban", ""))
        ws.append([idx, full_name, montant, bic, iban])

    last_data_row = ws.max_row
    total_row = last_data_row + 1
    ws.append(["", "Total", f"=SOMME(C9:C{last_data_row})", "", ""])
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    ws.cell(row=total_row, column=3).font = Font(bold=True)

    exchange_rate = 45
    mru_total_row = total_row + 1
    ws.append(["", "Total en MRU", f"=C{total_row}*{exchange_rate}", "", ""])
    ws.cell(row=mru_total_row, column=2).font = Font(bold=True)
    ws.cell(row=mru_total_row, column=3).font = Font(bold=True)

    ws.cell(row=mru_total_row + 1, column=4, value="Paris, le")

    sig_row = mru_total_row + 3
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=5)
    ws.cell(row=sig_row, column=1).value = "Mohamed Yahya Teiss                                               Abdellahi SALEM                                               Bettah Limam"
    ws.cell(row=sig_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=sig_row, column=1).font = Font(bold=True)

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Fichier Excel", "*.xlsx")],
        title="Exporter au format BOURSE"
    )
    if file_path:
        try:
            wb.save(file_path)
            messagebox.showinfo("Succès", f"Fichier exporté avec succès :\n{file_path}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur d'export : {e}")

def open_search():
    win = tk.Tk()
    win.title("Recherche")
    win.geometry("900x600")
    win.iconbitmap(resource_path("flag.ico"))
    win.resizable(False, False)

    entries = {}

    fields = [
        ("Nom", "nom"),
        ("Prenom", "prenom"),
        ("NNI", "nni"),
        ("Date de naissance", "dob"),
        ("Convention", "type_convention"),
        ("Année academique", "academic_year"),
        ("Domaine", "field_of_study"),
        ("Lieu de naissance", "place_of_birth"),
        ("Email", "email"),
        ("Telephone", "phone_number"),
        ("Ville d'études", "study_location")
    ]

    for i, (label_text, key) in enumerate(fields):
        tk.Label(win, text=label_text).grid(row=i, column=0, padx=10, pady=5, sticky="e")
        entry = tk.Entry(win, width=30)
        entry.grid(row=i, column=1, padx=10, pady=5)
        entries[key] = entry

    tk.Button(win, text="Recherche", command=lambda: recherche(entries)).grid(row=len(fields), column=0, columnspan=2, pady=20)

    win.mainloop()

if __name__ == "__main__":
    open_search()
